from openpyxl import Workbook

wb = Workbook()
sheet = wb.active
sheet["A1"] = "Hello"
sheet.cell(2, 1, "Goodbye")
cell = sheet.cell(3, 1)
# cell = sheet.cell(row=3, column=1)
cell.value = "Thank you"

for i, v in enumerate(["こんにちは", "さようなら", "ありがとう"]):
    sheet.cell(row=(i + 1), column=2, value=v)
    

wb.save("cell_write.xlsx")

from openpyxl import load_workbook

wb = load_workbook("uriage.xlsx", data_only=True)
ws = wb.active

# min_colは始まりの列番号、min_rowは始まりの行番号
rows  = ws.iter_rows(min_col=1, min_row=3)

for row in rows:
    values = [cell.value for cell in row]
    print(values)
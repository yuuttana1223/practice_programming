import datetime
from openpyxl import Workbook

wb = Workbook()
sheet = wb.active

this_year = datetime.date.today().year

for i in range(100 + 1):
    age = i
    year = this_year - age
    sheet.cell(row=i + 1, column=1, value=f"{age}歳")
    sheet.cell(row=i + 1, column=2, value=f"{year}年")
wb.save("agelist.xlsx")
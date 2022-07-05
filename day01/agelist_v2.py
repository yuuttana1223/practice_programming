import datetime
from openpyxl import Workbook

wb = Workbook()
sheet = wb.active

this_year = datetime.date.today().year
start_year = this_year - 20

row = 0
for year in range(start_year, this_year + 1):
    row += 1
    sheet.cell(row=row, column=1, value=f"{year}年度")
    sheet.cell(row=row, column=2, value=f"{year}年4月1日〜{year + 1}年3月31日")
wb.save("nendo.xlsx")
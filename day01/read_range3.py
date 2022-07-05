from openpyxl import load_workbook

wb = load_workbook("cellname100.xlsx")
sheet = wb.active

it  = sheet.iter_rows(min_row=2, max_row=4, min_col=3, max_col=6)

for row in it:
    values = [cell.value for cell in row]
    print(values)
    # ['C2', 'D2', 'E2', 'F2']
    # ['C3', 'D3', 'E3', 'F3']
    # ['C4', 'D4', 'E4', 'F4']
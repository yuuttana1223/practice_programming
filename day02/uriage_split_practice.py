from openpyxl import Workbook, load_workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font


book = load_workbook("uriage.xlsx", data_only=True)
sheet = book.active

rows = sheet.iter_rows(min_row=3, min_col=1)
column_names = ["NO", "納品日", "商品名", "単価", "数量", "金額", "担当営業"]
for i, row in enumerate(rows):
    values = [cell.value for cell in row]
    name = values[-1]
    if name in book.sheetnames:
        to_sheet = book[name]
    else:
        to_sheet = book.create_sheet(title=name)
        to_sheet["A1"].value = f"{name}の売上一覧"
        to_sheet.append(column_names)
        to_sheet.column_dimensions["c"].width = 20
        
    to_sheet.append(values)
    max_row = to_sheet.max_row
    to_sheet.cell(row=max_row, column=2).number_format = "m月d日"
    to_sheet.cell(row=max_row, column=4).number_format = "#,##0"
    to_sheet.cell(row=max_row, column=6).number_format = "#,##0"
    to_sheet["A1"].font = Font(name='メイリオ', size=14, bold=True)
    for row in to_sheet["A2" : "G2"]:
        for cell in row:
            cell.font = Font(name='メイリオ', size=12, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            top_side = Side(style='thin', color='000000')
            bottom_side = Side(style='double', color='000000')
            cell.border = Border(top=top_side, bottom=bottom_side)
book.save("uriage_split_practice.xlsx")
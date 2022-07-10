from openpyxl import Workbook, load_workbook

book = load_workbook("uriage.xlsx", data_only=True)
sheet = book.active

rows = sheet.iter_rows(min_row=3, min_col=1)
column_names = ["NO", "納品日", "商品名", "単価", "数量", "金額", "担当営業"]
for row in rows:
    values = [cell.value for cell in row]
    name = values[-1]
    if name in book.sheetnames:
        to_sheet = book[name]
    else:
        to_sheet = book.create_sheet(title=name)
        to_sheet["A1"].value = f"{name}の売上一覧"
        to_sheet.append(column_names)
        to_sheet.column_dimensions["c"].width = 20
        
    to_sheet.append(values)
    max_row = to_sheet.max_row
    to_sheet.cell(row=max_row, column=2).number_format = "m月d日"
    to_sheet.cell(row=max_row, column=4).number_format = "#,##0"
    to_sheet.cell(row=max_row, column=6).number_format = "#,##0"
    
book.save("uriage_split3.xlsx")
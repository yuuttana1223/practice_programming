from openpyxl import load_workbook

book = load_workbook("uriage.xlsx", data_only=True)
sheet = book.active

data = []
for row in sheet["A3" : "G28"]:
    values = [cell.value for cell in row]
    data.append(values)

data.sort(key=lambda x: (x[6], x[5]))

for row, row_val in enumerate(data):
    for col, col_val in enumerate(row_val):
        sheet.cell(row=row+3, column=col+1).value = col_val
        
book.save("uriage_sort2.xlsx")
from openpyxl import Workbook, load_workbook

read_book = load_workbook("uriage.xlsx", data_only=True)
read_sheet = read_book.worksheets[0]
rows = read_sheet.iter_rows(min_row=2,min_col=1)

# [{}, {}]
sales_li = []
column_names = []

for i, row in enumerate(rows):
    row_dict = {}
    for j, cell in enumerate(row):
        if i == 0:
            column_names.append(cell.value)
        else:
            row_dict[column_names[j]] = cell.value
    # 0にはカラムのタイトルが入っている
    if i != 0:
        sales_li.append(row_dict)


book = Workbook()
sheet = book.active
sheet["A1"].value = "佐藤の売上一覧"
for i, cell in enumerate(column_names):
    sheet.cell(row=2, column=i+1).value = cell

row_count = 2
for row in sales_li:
    if row["担当営業"] != "佐藤":
        continue
    row_count += 1
    for j, cell in enumerate(row.values()):
        sheet.cell(row=row_count, column=j+1).value = cell

book.save("uriage_sato.xlsx")
